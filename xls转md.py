import os
import tkinter as tk
from tkinter import filedialog, messagebox
from pathlib import Path
from typing import List, Dict, Any, Optional

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    import openpyxl
except ImportError:
    openpyxl = None


def install_dependencies():
    import subprocess
    import sys
    global xlrd, openpyxl
    packages = []
    if xlrd is None:
        packages.append('xlrd')
    if openpyxl is None:
        packages.append('openpyxl')
    if packages:
        print(f"正在安装缺少的依赖: {packages}")
        subprocess.check_call([sys.executable, '-m', 'pip', 'install'] + packages)
        import xlrd
        import openpyxl


def read_xls(file_path: str) -> List[List[str]]:
    workbook = xlrd.open_workbook(file_path)
    data = []
    for sheet in workbook.sheets():
        sheet_data = []
        for row_idx in range(sheet.nrows):
            row_data = []
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                row_data.append(str(cell.value))
            sheet_data.append(row_data)
        data.append(sheet_data)
    return data


def read_xlsx(file_path: str) -> List[List[str]]:
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    data = []
    for sheet in workbook.worksheets:
        sheet_data = []
        for row in sheet.iter_rows(values_only=True):
            row_data = [str(cell) if cell is not None else '' for cell in row]
            sheet_data.append(row_data)
        data.append(sheet_data)
    return data


def convert_to_markdown(data: List[List[str]], file_name: str) -> str:
    md_content = []
    md_content.append(f"# {file_name}\n")

    for sheet_idx, sheet_data in enumerate(data):
        if len(data) > 1:
            md_content.append(f"\n## Sheet {sheet_idx + 1}\n")

        if not sheet_data:
            continue

        col_count = len(sheet_data[0])
        md_content.append("| " + " | ".join([f"Column {i+1}" for i in range(col_count)]) + " |")
        md_content.append("| " + " | ".join(["---" for _ in range(col_count)]) + " |")

        for row in sheet_data:
            while len(row) < col_count:
                row.append('')
            md_content.append("| " + " | ".join(row) + " |")

    return "\n".join(md_content)


def select_xls_file() -> Optional[str]:
    file_path = filedialog.askopenfilename(
        title="选择Excel文件",
        filetypes=[
            ("Excel文件", "*.xls *.xlsx"),
            ("Excel 97-2003", "*.xls"),
            ("Excel 2007+", "*.xlsx"),
            ("所有文件", "*.*")
        ]
    )
    return file_path if file_path else None


def save_markdown(content: str, original_path: str):
    base_name = Path(original_path).stem
    default_filename = f"{base_name}.md"
    file_path = filedialog.asksaveasfilename(
        title="保存Markdown文件",
        defaultextension=".md",
        initialfile=default_filename,
        filetypes=[("Markdown文件", "*.md"), ("所有文件", "*.*")]
    )
    if file_path:
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(content)
        messagebox.showinfo("成功", f"已保存为: {file_path}")
        return True
    return False


def convert_xls_to_md(file_path: str = None) -> Optional[str]:
    if file_path is None:
        file_path = select_xls_file()

    if not file_path:
        return None

    if not os.path.exists(file_path):
        messagebox.showerror("错误", f"文件不存在: {file_path}")
        return None

    ext = Path(file_path).suffix.lower()

    try:
        if ext == '.xls':
            if xlrd is None:
                install_dependencies()
            data = read_xls(file_path)
        elif ext == '.xlsx':
            if openpyxl is None:
                install_dependencies()
            data = read_xlsx(file_path)
        else:
            messagebox.showerror("错误", "不支持的文件格式")
            return None

        file_name = Path(file_path).name
        md_content = convert_to_markdown(data, file_name)
        save_markdown(md_content, file_path)
        return md_content

    except Exception as e:
        messagebox.showerror("错误", f"转换失败: {str(e)}")
        return None


def main():
    root = tk.Tk()
    root.withdraw()

    convert_xls_to_md()


if __name__ == "__main__":
    main()
